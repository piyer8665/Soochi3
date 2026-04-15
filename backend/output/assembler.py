import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import SUMMARY_TABLE_MAX_CODE_LENGTH


SUMMARY_SIGNALS = {"whole", "total", "overall", "all", "other",
                   "unknown", "none", "missing", "combined"}


def _sort_coding_table(coding_table: list) -> list:
    if not coding_table:
        return coding_table

    # Deduplicate by normalized name before sorting
    # Merge frequency of duplicate names into the first occurrence
    seen_names = {}
    deduplicated = []
    for row in coding_table:
        name_norm = str(row.get("name", "")).lower().strip()
        # Strip common suffixes like "(alternate entry)", "(instance 1)" etc.
        import re
        name_clean = re.sub(r'\s*\(.*?\)\s*$', '', name_norm).strip()
        if name_clean in seen_names:
            # Merge frequency into existing entry
            existing = deduplicated[seen_names[name_clean]]
            existing["frequency"] = existing.get("frequency", 0) + row.get("frequency", 0)
        else:
            seen_names[name_clean] = len(deduplicated)
            deduplicated.append(dict(row))

    def sort_key(row):
        name = str(row.get("name", "")).lower().strip()
        words = set(name.split())
        is_summary = bool(words & SUMMARY_SIGNALS)
        return (1 if is_summary else 0, name)

    sorted_table = sorted(deduplicated, key=sort_key)
    for idx, row in enumerate(sorted_table, 1):
        row["code"] = str(idx)
    return sorted_table


def assemble_dictionary(deterministic_entries: list, writer_entries: list,
                         normality_report=None) -> dict:
    all_entries = {}

    for entry in writer_entries:
        col = entry.get("column", "").strip()
        if col:
            all_entries[col] = entry
    # Deterministic entries override writer — deterministic always wins
    for entry in deterministic_entries:
        col = entry.get("column", "").strip()
        if col:
            all_entries[col] = entry
            if col in ['hypertension', 'heart_disease', 'stroke']:
                import sys
                codes = [(r.get('code'), r.get('name')) for r in entry.get('coding_table', [])]
                print(f'[ASSEMBLER DEBUG] {col} stored: {codes}', file=sys.stderr)


    # Two-pass: Python sorts coding tables alphabetically by name
    # UNLESS the Interpreter signals domain_convention or user_specified ordering
    for col, entry in all_entries.items():
        if entry.get("coding_table"):
            ordering_basis = entry.get("ordering_basis", "alphabetical")
            # Binary 0/1 variables — preserve codes exactly, no renumbering
            if ordering_basis == "binary_zero_one":
                pass
            else:
                subtypes = entry.get("subtypes", [])
                is_uncoded = "needs_coding" in subtypes or not subtypes
                preserve_order = (
                    not is_uncoded and
                    ordering_basis in ("domain_convention", "user_specified")
                )
                if preserve_order:
                    # Preserve Interpreter ordering — just reassign codes 1-N
                    for idx, row in enumerate(entry["coding_table"], 1):
                        row["code"] = str(idx)
                else:
                    entry["coding_table"] = _sort_coding_table(entry["coding_table"])

    # Post-process binary yes/no text columns — align dictionary codes with recoder 0/1 output
    neg_words = {"no", "false", "absent", "negative", "never", "without", "non"}
    pos_words = {"yes", "true", "present", "positive", "always", "with"}
    for col, entry in all_entries.items():
        ct = entry.get("coding_table", [])
        if len(ct) == 2 and entry.get("variable_type") in ("Categorical Nominal", "Categorical Ordinal"):
            names = [str(r.get("name", "")).lower().strip() for r in ct]
            has_neg = any(n in neg_words for n in names)
            has_pos = any(n in pos_words for n in names)
            if has_neg and has_pos:
                for row in ct:
                    name_lower = str(row.get("name", "")).lower().strip()
                    if name_lower in neg_words:
                        row["code"] = "0"
                    elif name_lower in pos_words:
                        row["code"] = "1"

    if normality_report:
        for result in normality_report.results:
            col = result.column.strip()
            if col in all_entries:
                notes = all_entries[col].get("data_quality_notes", [])
                notes.append(f"Normality: {result.test} p={result.p_value} — {result.recommendation}")
                all_entries[col]["data_quality_notes"] = notes
                all_entries[col]["normality_passes"] = result.passes
                all_entries[col]["normality_recommendation"] = result.recommendation

    return all_entries


def format_dictionary_text(entries: dict, dataset_name: str,
                            total_rows: int, total_columns: int) -> str:
    lines = []
    lines.append(f"# Data Dictionary: {dataset_name}")
    lines.append(f"## Dataset: {total_rows} rows, {total_columns} variables")
    lines.append("")

    for col, entry in entries.items():
        var_type = entry.get("variable_type", "Unknown")
        description = entry.get("description", "")
        coding_table = entry.get("coding_table", [])
        data_quality_notes = entry.get("data_quality_notes", [])
        range_val = entry.get("range")
        normality_rec = entry.get("normality_recommendation")

        lines.append(f"## {col}")
        lines.append(f"**Type:** {var_type}")
        lines.append(f"**Description:** {description}")

        if range_val:
            lines.append(f"**Range:** {range_val}")
            if entry.get("mean") is not None:
                lines.append(f"**Mean:** {entry.get('mean')} | **Median:** {entry.get('median')} | **SD:** {entry.get('std')}")

        if normality_rec:
            lines.append(f"**Statistical assumption:** {normality_rec}")

        if coding_table:
            lines.append("")
            lines.append("### Coding Table")
            lines.append("| Code | Name | Definition |")
            lines.append("|------|------|------------|")
            for row in coding_table:
                code = str(row.get("code", ""))
                name = str(row.get("name", ""))
                definition = str(row.get("definition", ""))
                lines.append(f"| {code} | {name} | {definition} |")

        if data_quality_notes:
            lines.append("")
            lines.append("**Data Quality Notes:**")
            for note in data_quality_notes:
                if not note.startswith("Normality:"):
                    lines.append(f"- {note}")

        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append("## Variable Summary")
    lines.append("")
    lines.append("| Variable | Type | Coded | Codes |")
    lines.append("|----------|------|-------|-------|")

    for col, entry in entries.items():
        var_type = entry.get("variable_type", "Unknown")
        coding_table = entry.get("coding_table", [])
        coded = "Yes" if coding_table else "No"
        if coding_table:
            codes_str = ", ".join([
                f"{r.get('code')}={r.get('name', '')}"
                for r in coding_table[:5]
            ])
            if len(codes_str) > SUMMARY_TABLE_MAX_CODE_LENGTH:
                codes_str = codes_str[:SUMMARY_TABLE_MAX_CODE_LENGTH] + "..."
        else:
            codes_str = "—"
        lines.append(f"| {col} | {var_type} | {coded} | {codes_str} |")

    return "\n".join(lines)


def build_excel_summary(entries: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variable Summary"

    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Variable", "Type", "Description", "Coded", "Sample Codes",
               "Range", "Mean", "SD", "Normality"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_idx, (col, entry) in enumerate(entries.items(), 2):
        coding_table = entry.get("coding_table", [])
        sample_codes = ", ".join([
            f"{r.get('code')}={r.get('name', '')}"
            for r in coding_table
        ]) if coding_table else "—"

        normality = "Pass" if entry.get("normality_passes") is True else (
            "Fail" if entry.get("normality_passes") is False else "—"
        )

        values = [
            col,
            entry.get("variable_type", ""),
            entry.get("description", "")[:200],
            "Yes" if coding_table else "No",
            sample_codes,
            entry.get("range", "—") or "—",
            str(entry.get("mean", "—")) if entry.get("mean") is not None else "—",
            str(entry.get("std", "—")) if entry.get("std") is not None else "—",
            normality
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    col_widths = [20, 20, 50, 8, 40, 15, 10, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer