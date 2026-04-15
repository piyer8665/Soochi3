"""
output/recoder.py
Soochi 2.0 — Recoded Dataset Generator

Produces a single Excel workbook with three sheets:
  Sheet 1: Uncoded Data  — original raw data as uploaded
  Sheet 2: Coded Data    — recoded dataset with numeric codes replacing text values
  Sheet 3: Data Dictionary — full coding tables and variable definitions
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def build_recoded_dataset(df: pd.DataFrame, entries: dict) -> io.BytesIO:
    """
    Main entry point.
    Takes original DataFrame and assembled dictionary entries.
    Returns BytesIO Excel buffer with three sheets.
    """
    recoded_df = _recode_dataframe(df, entries)
    buffer = _build_excel(df, recoded_df, entries)
    return buffer


def _recode_dataframe(df: pd.DataFrame, entries: dict) -> pd.DataFrame:
    """
    Builds lookup tables from coding tables and applies recoding.
    Returns recoded DataFrame.
    """
    recoded_df = df.copy()

    for col_clean, entry in entries.items():
        variable_type = entry.get("variable_type", "")
        coding_table = entry.get("coding_table", [])

        # Find actual column in DataFrame (handles trailing spaces)
        actual_col = col_clean if col_clean in df.columns else next(
            (c for c in df.columns if c.strip() == col_clean.strip()), None
        )

        if not actual_col:
            continue

        # Only recode text categorical variables with a coding table
        if variable_type not in ("Categorical Nominal", "Categorical Ordinal"):
            continue

        if not coding_table:
            continue

        series = df[actual_col]

        # Handle binary 0/1 columns that represent Y/N or similar
        # Skip all numeric columns — leave as-is, no remapping
        if pd.api.types.is_numeric_dtype(series):
            continue

        # Build lookup: normalized original text value -> numeric code
        text_lookup = {}
        original_values = series.dropna().unique()

        for orig_val in original_values:
            orig_str = str(orig_val).strip()
            orig_norm = orig_str.lower()
            best_match = None
            best_score = 0.0
            for row in coding_table:
                code_name = str(row.get('name', '')).strip().lower()
                code_num = row.get('code', '')
                # Direct normalized match
                if orig_norm == code_name:
                    best_match = code_num
                    best_score = 1.0
                    break
                # Partial match
                if orig_norm in code_name or code_name in orig_norm:
                    if best_score < 0.9:
                        best_match = code_num
                        best_score = 0.9
                # Fuzzy match for typo variants
                score = _similarity(orig_norm, code_name)
                if score >= 0.8 and score > best_score:
                    best_match = code_num
                    best_score = score
            if best_match is not None:
                try:
                    text_lookup[orig_str] = int(best_match)
                except (ValueError, TypeError):
                    text_lookup[orig_str] = best_match

        # For binary yes/no text columns, remap to 0/1 industry standard
        neg_words = {'no', 'false', 'absent', 'negative', 'never'}
        pos_words = {'yes', 'true', 'present', 'positive', 'always'}
        if len(text_lookup) == 2:
            keys = list(text_lookup.keys())
            key_norms = [k.lower().strip() for k in keys]
            has_neg = any(k in neg_words for k in key_norms)
            has_pos = any(k in pos_words for k in key_norms)
            if has_neg and has_pos:
                for k in list(text_lookup.keys()):
                    if k.lower().strip() in neg_words:
                        text_lookup[k] = 0
                        text_lookup[str(k).strip()] = 0
                    else:
                        text_lookup[k] = 1
                        text_lookup[str(k).strip()] = 1

        if not text_lookup:
            continue

        def recode_value(val):
            if pd.isna(val):
                return val
            val_str = str(val)
            # Try exact match first, then stripped match
            return text_lookup.get(val_str, text_lookup.get(val_str.strip(), val))

        recoded_df[actual_col] = series.apply(recode_value)

    return recoded_df


def _build_excel(original_df: pd.DataFrame, recoded_df: pd.DataFrame, entries: dict) -> io.BytesIO:
    """Builds the three-sheet Excel workbook."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    def write_dataframe(ws, df):
        """Write a DataFrame to a worksheet with header formatting."""
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_top

        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top")

        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(12, min(40, len(str(col_name)) + 4))

    # ── Sheet 1: Uncoded Data ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Uncoded Data"
    write_dataframe(ws1, original_df)

    # ── Sheet 2: Coded Data ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Coded Data")
    write_dataframe(ws2, recoded_df)

    # ── Sheet 3: Data Dictionary ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Data Dictionary")

    dict_headers = ["Variable", "Type", "Code", "Code Name", "Definition", "Range", "Data Quality Notes"]
    for col_idx, header in enumerate(dict_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_top

    row_idx = 2
    for col_name, entry in entries.items():
        variable_type = entry.get("variable_type", "")
        coding_table = entry.get("coding_table", [])
        range_val = entry.get("range", "—") or "—"
        notes = "; ".join(entry.get("data_quality_notes", [])[:3]) or "—"

        if coding_table:
            for code_row in coding_table:
                ws3.cell(row=row_idx, column=1, value=col_name).alignment = wrap_top
                ws3.cell(row=row_idx, column=2, value=variable_type).alignment = wrap_top
                ws3.cell(row=row_idx, column=3, value=code_row.get("code", "")).alignment = wrap_top
                ws3.cell(row=row_idx, column=4, value=code_row.get("name", "")).alignment = wrap_top
                ws3.cell(row=row_idx, column=5, value=code_row.get("definition", "")).alignment = wrap_top
                ws3.cell(row=row_idx, column=6, value=range_val).alignment = wrap_top
                ws3.cell(row=row_idx, column=7, value=notes if row_idx == 2 or ws3.cell(row=row_idx-1, column=1).value != col_name else "").alignment = wrap_top
                row_idx += 1
        else:
            ws3.cell(row=row_idx, column=1, value=col_name).alignment = wrap_top
            ws3.cell(row=row_idx, column=2, value=variable_type).alignment = wrap_top
            ws3.cell(row=row_idx, column=3, value="—").alignment = wrap_top
            ws3.cell(row=row_idx, column=4, value="—").alignment = wrap_top
            ws3.cell(row=row_idx, column=5, value=entry.get("description", "")).alignment = wrap_top
            ws3.cell(row=row_idx, column=6, value=range_val).alignment = wrap_top
            ws3.cell(row=row_idx, column=7, value=notes).alignment = wrap_top
            row_idx += 1

    for col_idx, width in enumerate([25, 20, 8, 25, 50, 15, 40], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
